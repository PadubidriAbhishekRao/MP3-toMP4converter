import fitz  # PyMuPDF
import openpyxl
from openpyxl import Workbook

def extract_comments_to_excel(pdf_path, excel_path):
    # Open the PDF file
    pdf_document = fitz.open(pdf_path)
    
    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PDF Comments"
    
    # Add headers to the Excel sheet
    sheet["A1"] = "Page Number"
    sheet["B1"] = "Comment Type"
    sheet["C1"] = "Comment Text"
    
    row = 2  # Start from the second row (after headers)
    
    # Iterate through each page in the PDF
    for page_num in range(len(pdf_document)):
        page = pdf_document[page_num]
        
        # Extract comments (annotations) from the page
        annotations = page.annots()
        
        if annotations:
            for annot in annotations:
                if annot.type[1] in ['Text', 'FreeText', 'Highlight', 'Underline', 'StrikeOut', 'Square', 'Circle', 'Line', 'Polygon', 'PolyLine', 'Stamp', 'Caret', 'Ink', 'FileAttachment']:
                    # Extract comment information
                    comment_type = annot.type[1]
                    comment_text = annot.info.get("content", "")
                    
                    # Write to Excel
                    sheet.cell(row=row, column=1, value=page_num + 1)
                    sheet.cell(row=row, column=2, value=comment_type)
                    sheet.cell(row=row, column=3, value=comment_text)
                    
                    row += 1
    
    # Save the Excel file
    workbook.save(excel_path)
    print(f"Comments extracted and saved to {excel_path}")

# Usage example
pdf_file = "C:/Users/HP/Documents/my_documents.pdf"
excel_file = "C:/Users/HP/Documents/hello.xlsx"
extract_comments_to_excel(pdf_file, excel_file)